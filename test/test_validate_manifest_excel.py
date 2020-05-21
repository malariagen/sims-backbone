from APPS.register_samples.errors import FieldError

from APPS.register_samples.validators import BulkAddSamplesValidatorFactory
from APPS.bulk_upload.errors import HeaderError, ValidationError
import APPS.register_samples.helpers as helpers
from APPS.bulk_upload.excel_parser import BulkUploadExcelConfig
import os

import datetime
import json
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.cell.cell import TYPE_FORMULA
import tempfile
import string
import uuid

import pytest

import openapi_client

from openapi_server.models.locations import Locations  # noqa: E501
from openapi_server.models.location import Location  # noqa: E501

from backbone_server.model.original_sample import OriginalSample
from backbone_server.model.attr import Attr
from backbone_server.model.study import Study

from test_base import TestBase

ASSETS_DIR = os.path.dirname(os.path.realpath(__file__))
MNF_XLS_SHEETNAME_SAMPLE = "Sample"

MNF_HEADER_DB_DISPLAY_MAP = {"external_id": "Specimen ID",
                             "collection_date": "Collection Date",
                             "location_id": "Collection Location (see locations sheet)",
                             "note": "Notes"}

MNF_DIR = os.path.dirname(os.path.realpath(__file__)) + \
    os.sep + os.pardir + os.sep + "assets"


class TestValidateManifestExcel(TestBase):
    """
    Testcases for validating excel manifest
    """

    @classmethod
    def setup_class(cls):
        """ setup any state specific to the execution of the given class (which
        usually contains tests).
        """
        location = Location(location_id=uuid.uuid4(),
                            curated_name='Bougoula-hameau', latitude="17.5",
                            longitude="-3.99")
        location2 = Location(location_id=uuid.uuid4(), curated_name='CSV Test Location2')
        locations = Locations()
        locations.locations = [location]
        locations.count = 1
        TestValidateManifestExcel.location = location
        TestValidateManifestExcel.location2 = location2
        study = Study(code='8888', name='8888-VAL-MD_UP')
        study.locations = locations
        TestValidateManifestExcel.study = study
        TestValidateManifestExcel.alt_study = Study(code='8887', name='8887-VAL-MD_UP')
        TestValidateManifestExcel.study2 = Study(code='8886', name='8886-VAL-MD_UP')



    @classmethod
    def teardown_class(cls):
        """ teardown any state that was previously setup with a call to
        setup_class.
        """
        pass

    registered_samples = []
    test_study = None

    def lookup_sample_func(self, field, value):

        for samp in self.registered_samples:
            if samp[field] == value:
                ret = OriginalSample()
                ret.attrs = [Attr(attr_type=field, attr_value_str=value)]
                ret.study = self.test_study
                return ret

        return None

    def test_validate_manifest_can_convert_dates(self):
        """
        Feature: XLS Manifest Date Autoconversion
            Roma should handle all the date conversion for you as long as the XLS cells are Date type.
            Whether the XLSX was created on MAC, Windows, and Linux where the epoch dates are different should not matter.
            Whether the XLSX was created on a region  locale different from the server should not matter.

            NB:  due to excel inability to handle dates before 1900, and the fact that
                samples will not likely have been collected earlier than this, ROMA will not handle this edge case.
            http://www.exceluser.com/formulas/earlydates.htm

            NB:  this is more of a test of openpyxl, but it allows us to know if backwards compatibility issues arise if we switch out packages
        """

        # The same input manifest XLS tests multiple scenarios.

        # SCENARIO: Manifest XLS Collection Date should not return formatting errors as long as it uses Date Type for any date format.
        #     Day and month format should not be dependent on region.  EG)  Jan 5, 2016 displayed as 1/5/2016 in US should not be confused with May 1, 2016 in UK.
        # GIVEN we validate manifest XLS with Collection Date cells set as Date Type with date format different from that used by Django
        # AND collection date cells contain a day <= 12
        # THEN no formatting errors occur and month, day, year are parsed correctly
        # AND manifest header ini config file is parsed correctly to map db column names to excel sheet column names

        # SCENARIO: Manifest XLS has last active row as last sample row.
        # GIVEN we validate manifest XLS where last sample row is last active row
        # THEN all sample rows are returned and parsed correctly
        # AND manifest header ini config file is parsed correctly to map db column names to excel sheet column names

        # SCENARIO: Manifest XLS created on MAC.
        # GIVEN we validate manifest XLS created on MAC
        # AND collection date cells contain a day <= 12
        # THEN no formatting errors occur and month, day, year are parsed correctly
        # AND manifest header ini config file is parsed correctly to map db column names to excel sheet column names

        # This spreadsheet uses a different location than was setup in create_study()

        workbook_fname = ASSETS_DIR + os.sep + \
            "SpotMalaria_TestManifest_CollectionDate_DateType_WrongFormat_Mac.xlsx"
        validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                    manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                    study=self.study, sample_lookup_func=self.lookup_sample_func)
        samples_iter = validator.validate_manifest_file()

        actual_data = []
        for resp in samples_iter:
            actual_data.append(resp["sample"])
        # NB:  if there is no value for the columns, then it won't be included in the return dict
        expected_data = [{"external_id": "Specimen1",
                          "location_id": self.location.location_id,
                          "collection_date": datetime.date(2015, 3, 12)},
                         {"external_id": "Specimen2",
                             "location_id": self.location.location_id,
                             "collection_date": datetime.date(1920, 1, 18)}]

        assert actual_data == expected_data

        # SCENARIO: Manifest XLS created on Windows.
        # GIVEN we validate manifest XLS created on Windows
        # AND collection date cells contain a day <= 12
        # THEN no formatting errors occur and month, day, year are parsed correctly
        # AND manifest header ini config file is parsed correctly to map db column names to excel sheet column names

        # SCENARIO: Manifest XLS Collection Date should not return formatting errors as long as it uses Date Type for any date display format.
        #     Display format should not matter
        # EG)  If date was formatted with *DD/MM/YYYY in excel, which indicates that the display will autochange depending on the viewer's region,
        # GIVEN we validate manifest XLS with Collection Date cells set as Date Type with date format set to autochange depending on excel viewer's region
        # AND collection date cells contain a day <= 12
        # THEN no formatting errors occur and month, day, year are parsed correctly
        # AND manifest header ini config file is parsed correctly to map db column names to excel sheet column names

        # SCENARIO: Manifest XLS has last active row past last sample row.
        # GIVEN we validate manifest XLS where last sample row is past last sample row
        # THEN all sample rows are returned and parsed correctly
        # AND there are no empty sample rows returned.

        workbook_fname = ASSETS_DIR + os.sep + \
            "SpotMalaria_TestManifest_CollectionDate_DateType_WrongFormat_Win.xlsx"
        validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                    manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                    study=self.study, sample_lookup_func=self.lookup_sample_func)
        samples_iter = validator.validate_manifest_file()

        actual_data = []
        for resp in samples_iter:
            actual_data.append(resp["sample"])
        # NB:  if there is no value for the columns, then it won't be included in the return dict
        expected_data = [{"external_id": "SpecimenIDWin1",
                          "location_id": self.location.location_id,
                          "collection_date": datetime.date(2015, 12, 3),
                          "note": "test note1"},
                         {"external_id": "SpecimenIDWin2",
                             "location_id": self.location.location_id,
                             "collection_date": datetime.date(1900, 12, 3),
                             "note": "test note2\nnewline\nanother line"},  # right and left strip
                         {"external_id": "SpecimenIDWin3",
                             "location_id": self.location.location_id,
                             "collection_date": datetime.date(2015, 12, 18),
                             "note": "testnote3  space"}
                         ]

        assert actual_data == expected_data

    def test_blank_rows(self):
        """
        GIVEN:  I upload a sample xlsx containing a valid Samples worksheet with blank rows in between data rows
        WHEN:  we validate the worksheet
        THEN: the blank rows are stripped out
        AND: the samples are validated successfully

        """
        workbook_fname = None
        try:
            TOTAL_ROWS = 5
            TOTAL_NONEMPTY_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            # openpyxl uses 1-based indexing, but python uses 0-based indexing
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            # First fill in all of the rows with data except one.  Then append blank rows. Then add the last data row.
            for irow in range(TOTAL_NONEMPTY_ROWS - 1):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            for irow in range(TOTAL_NONEMPTY_ROWS - 1, TOTAL_ROWS - 1):
                ws.append((None, None, None, None))

            external_id_cell = Cell(
                ws, value="Sample" + str(TOTAL_NONEMPTY_ROWS - 1))
            date_cell = Cell(ws, value=datetime.date(
                year=2000, month=1, day=19))
            date_cell.number_format = "yyyy-mm-dd"
            loc_cell = Cell(ws, value=self.location.curated_name)
            note_cell = Cell(ws, value="")
            last_row = (external_id_cell, date_cell, loc_cell, note_cell)
            ws.append(last_row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="blankRows_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            samples_iter = validator.validate_manifest_file()

            actual_data = []
            for resp in samples_iter:
                actual_data.append(resp["sample"])
            # NB:  if there is no value for the columns, then it won't be included in the return dict
            expected_data = [{"external_id": "Sample0",
                              "location_id": self.location.location_id,
                              "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample1",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample2",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)}
                             ]

            assert actual_data == expected_data

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_missing_header(self):
        """
        GIVEN:  I upload a sample xlsx containing a Samples worksheet with a missing required header
        WHEN:  we validate the worksheet
        THEN: a ValidationError is thrown

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            # Missing collection_date header - wrongly named, non-string cell values
            ws.append(
                (db_to_nice_header["external_id"], "Dollection Date", 5, datetime.datetime.now()))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="missingHeader_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)

            with pytest.raises(ValidationError) as e:
                samples_iter = validator.validate_manifest_file()
                for sample_dict in samples_iter:
                    if sample_dict and len(sample_dict):
                        pass

            assert e.value.code == HeaderError.MISSING_HEADER
            assert all(c in string.printable for c in str(e.value)), "The Missing Header Validation Error has not been properly interpolated"

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_blank_column(self):
        """
        GIVEN:  I upload a sample xlsx containing a valid Samples worksheet with a blank column in between required columns
        WHEN:  we validate the worksheet
        THEN: the blank columns are stripped out
        AND: the samples are validated successfully

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            # Missing collection_date header - wrongly named
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       "", db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                empty_cell = Cell(ws, value="")
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell,
                       empty_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="blankCol_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            samples_iter = validator.validate_manifest_file()

            actual_data = []
            for resp in samples_iter:
                actual_data.append(resp["sample"])
            # NB:  if there is no value for the columns, then it won't be included in the return dict
            expected_data = [{"external_id": "Sample0",
                              "location_id": self.location.location_id,
                              "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample1",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample2",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)}
                             ]

            assert actual_data == expected_data

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_invalid_loc(self):
        """
        GIVEN:  I upload a sample xlsx containing a Samples worksheet a location name that is not a valid location
        WHEN:  we validate the worksheet
        THEN: we get a ValidationError with an InvalidField code

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(
                    ws, value="Made up location that doesn't exist anywhere")
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="wrongLoc_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            total_samples = 0
            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            with pytest.raises(ValidationError) as e:

                for sample_dict in validator.validate_manifest_file():
                    if sample_dict and len(sample_dict):
                        total_samples += 1
            assert e.value.code == FieldError.INVALID_LOC, "Expected " + str(FieldError.INVALID_LOC) + " exception for invalid location but got " + str(e.value.code)
            assert all(c in string.printable for c in str(e.value)), "The Invalid Field Validation Error has not been properly interpolated"
            assert total_samples == 0, "Expected no samples but got " + str(total_samples)

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_invalid_loc_multiple(self):
        """
        GIVEN:  I upload a sample xlsx containing a Samples worksheet a location name that is not a valid location
        WHEN:  we validate the worksheet
        THEN: we get a ValidationError with an InvalidField code

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(
                    ws, value="Made up location that doesn't exist anywhere")
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="wrongLoc_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            total_samples = 0
            total_errors = 0
            errors = []

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func,
                                                                        is_raise_field_err=False)

            for sample_dict in validator.validate_manifest_file():
                if sample_dict["valid"]:
                    if sample_dict and len(sample_dict):
                        total_samples += 1
                else:
                    total_errors += 1
                    errors += sample_dict["errors"]

            assert total_errors == 3

            for idx, val in enumerate(errors):
                err_msg = "Failed validation for coord = C{}, field='location_id', value='Made up location that doesn't exist anywhere' [\"Made up location that doesn't exist anywhere is not a location registered with study project code {}. If you want to add this location, please contact the Coordinator.\"]".format(idx + 2, self.study.name)
                assert val == err_msg

            assert total_samples == 0, "Expected no samples but got " + str(total_samples)

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

        def test_invalid_date(self):
            """
            GIVEN:  I upload a sample xlsx containing a Samples worksheet with a collection date that is in wrong format
            WHEN:  we validate the worksheet
            THEN: we get a ValidationError with an InvalidField code

            """
            workbook_fname = None
            try:

                TOTAL_ROWS = 3

                manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
                manifest_config = BulkUploadExcelConfig(
                    header_config_file=manifest_config_ini)
                nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
                config_sheet_name = manifest_config.get_config_sheet_name()

                # There appears to be a difference in functionality of write only vs
                # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
                wb = Workbook()
                ws = wb.create_sheet(title=config_sheet_name)

                # Set the header
                ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                           db_to_nice_header["location_id"], db_to_nice_header["note"]))

                # Fill in the rows
                for irow in range(TOTAL_ROWS):
                    external_id_cell = Cell(ws, value="Sample" + str(irow))
                    # If we keep this as datetime.date, then openpyxl will convert to Date excel format, and display format doesn't matter
                    # We want string so that the way it is displayed affects how its parsed.
                    date_cell = Cell(ws, value=datetime.date(
                        year=2000, month=1, day=19).strftime("%Y"))
                    loc_cell = Cell(
                        ws, value="Made up location that doesn't exist anywhere")
                    note_cell = Cell(ws, value=None)
                    row = (external_id_cell, date_cell, loc_cell, note_cell)

                    ws.append(row)

                tmp = tempfile.NamedTemporaryFile(
                    suffix=".xlsx", prefix="wrongDate_SampleManifest", delete=False)
                tmp.close()
                workbook_fname = tmp.name
                wb.save(tmp.name)

                total_samples = 0
                validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                            manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                            study=self.study, sample_lookup_func=self.lookup_sample_func)
                with pytest.raises(ValidationError) as e:

                    for sample_dict in validator.validate_manifest_file():
                        if sample_dict and len(sample_dict):
                            total_samples += 1
                assert e.value.code == FieldError.INVALID_DATE, "Expected invalid field exception for invalid date but got " + str(e.value.code)
                assert all(c in string.printable for c in str(e.value)), "The Invalid Field Validation Error has not been properly interpolated"
                assert total_samples == 0, "Expected no samples but got " + str(total_samples)

            finally:
                if workbook_fname and os.path.exists(workbook_fname):
                    print("Deleting test file " + str(workbook_fname))
                    os.remove(workbook_fname)

        def test_invalid_future_date(self):
            """
            GIVEN:  I upload a sample xlsx containing a Samples worksheet with
            a collection date that is in the future
            WHEN:  we validate the worksheet
            THEN: we get a ValidationError with an InvalidField code

            """
            workbook_fname = None
            try:

                TOTAL_ROWS = 3

                manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
                manifest_config = BulkUploadExcelConfig(
                    header_config_file=manifest_config_ini)
                nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
                config_sheet_name = manifest_config.get_config_sheet_name()

                # There appears to be a difference in functionality of write only vs
                # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
                wb = Workbook()
                ws = wb.create_sheet(title=config_sheet_name)

                # Set the header
                ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                           db_to_nice_header["location_id"], db_to_nice_header["note"]))

                # Fill in the rows
                for irow in range(TOTAL_ROWS):
                    external_id_cell = Cell(ws, value="Sample" + str(irow))
                    date_cell = Cell(ws, value=datetime.date(year=2100, month=1, day=19))
                    loc_cell = Cell(
                        ws, value="Made up location that doesn't exist anywhere")
                    note_cell = Cell(ws, value=None)
                    row = (external_id_cell, date_cell, loc_cell, note_cell)

                    ws.append(row)

                tmp = tempfile.NamedTemporaryFile(
                    suffix=".xlsx", prefix="wrongDate_SampleManifest", delete=False)
                tmp.close()
                workbook_fname = tmp.name
                wb.save(tmp.name)

                total_samples = 0
                validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                            manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                            study=self.study, sample_lookup_func=self.lookup_sample_func)
                with pytest.raises(ValidationError) as e:

                    for sample_dict in validator.validate_manifest_file():
                        if sample_dict and len(sample_dict):
                            total_samples += 1
                assert e.value.code == FieldError.INVALID_DATE, "Expected invalid field exception for invalid date but got " + str(e.value.code)
                assert all(c in string.printable for c in str(e.value)), "The Invalid Field Validation Error has not been properly interpolated"
                assert total_samples == 0, "Expected no samples but got " + str(total_samples)

            finally:
                if workbook_fname and os.path.exists(workbook_fname):
                    print("Deleting test file " + str(workbook_fname))
                    os.remove(workbook_fname)

    def test_missing_samples_worksheet(self):
        """
        GIVEN:  I upload a sample xlsx containing a Samples worksheet a location name that is not a valid location
        WHEN:  we validate the worksheet
        THEN: we get a ValidationError with an InvalidField code

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=str(datetime.datetime.now()).replace(
                "-", "").replace(":", "").replace(".", ""))

            # Set the header
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="missingSheet_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            total_samples = 0
            expected_valid_loc = {self.location.curated_name: 1}
            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            with pytest.raises(ValidationError) as e:

                for sample_dict in validator.validate_manifest_file():
                    if sample_dict and len(sample_dict):
                        total_samples += 1
            assert e.value.code == FieldError.MISSING_SAMPLES_TAB, "Expected Missing Samples Sheet exception but got " + str(e.value.code)
            assert all(c in string.printable for c in str(e.value)), "The Missing Samples Sheet Validation Error has not been properly interpolated"
            assert total_samples == 0, "Expected no samples but got " + str(total_samples)

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_case_insensitive_lead_trail_space_loc(self):
        """
        GIVEN:  I upload a sample xlsx containing a valid Samples worksheet
        AND:  sample location name uses a different case than stored in database
        AND:  sample location name uses trailing and leading spaces
        WHEN:  we validate the worksheet
        THEN: samples validate successfully

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(
                    ws, value=" " + self.location.curated_name.upper() + "\n")
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="caseInsensitiveTrailingLeadSpaceLoc_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            expected_valid_loc = {" phnom penh\n": 1}
            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            samples_iter = validator.validate_manifest_file()

            actual_data = []
            for resp in samples_iter:
                actual_data.append(resp["sample"])
            expected_data = [{"external_id": "Sample0",
                              "location_id": self.location.location_id,
                              "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample1",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample2",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)}
                             ]

            assert actual_data == expected_data

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_optional_col(self):
        """
        GIVEN:  I upload a sample xlsx containing a valid Samples worksheet with non-mandatory columns
        WHEN:  we validate the worksheet
        THEN: all non-mandatory columns are combined into a single json object as the validated value

        """
        workbook_fname = None
        try:
            TOTAL_NONEMPTY_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            # openpyxl uses 1-based indexing, but python uses 0-based indexing
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"], db_to_nice_header["location_id"], db_to_nice_header["note"],
                       "PatientID", "Days From Baseline"))

            # Fill in the rows
            # First fill in all of the rows with data except one.  Then append blank rows. Then add the last data row.
            for irow in range(TOTAL_NONEMPTY_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                patient_cell = Cell(ws, value="Pat" + str(irow))
                d_baseline_cell = Cell(ws, value=irow)
                row = (external_id_cell, date_cell, loc_cell,
                       note_cell, patient_cell, d_baseline_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="optionalCol_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        #                                                                     manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            samples_iter = validator.validate_manifest_file()

            actual_data = []
            for resp in samples_iter:
                actual_data.append(resp["sample"])
            # NB:  if there is no value for the columns, then it won't be included in the return dict
            expected_data = [{"external_id": "Sample0",
                              "location_id": self.location.location_id,
                              "collection_date": datetime.date(year=2000, month=1, day=19),
                              "tags": {"PatientID": "Pat0",
                                       "Days From Baseline": 0
                                       }
                              },
                             {"external_id": "Sample1",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19),
                                 "tags": {"PatientID": "Pat1",
                                          "Days From Baseline": 1
                                          }
                              },
                             {"external_id": "Sample2",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19),
                                 "tags": {"PatientID": "Pat2",
                                          "Days From Baseline": 2
                                          }
                              }
                             ]

            assert actual_data == expected_data

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_blank_header_nonblank_cell(self):
        """
        GIVEN:  I upload a sample xlsx containing a valid Samples worksheet with a column with non-blank cell values but blank header
        WHEN:  we validate the worksheet
        THEN: the column with blank header is ignored

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            # Missing collection_date header - wrongly named
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       "", db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                nonempty_cell = Cell(ws, value="non-blank")
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell,
                       nonempty_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="blankCol_nonBlankCell_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            samples_iter = validator.validate_manifest_file()

            actual_data = []
            for resp in samples_iter:
                actual_data.append(resp["sample"])
            # NB:  if there is no value for the columns, then it won't be included in the return dict
            expected_data = [{"external_id": "Sample0",
                              "location_id": self.location.location_id,
                              "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample1",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample2",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)}
                             ]

            assert actual_data == expected_data

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_dup_header(self):
        """
        GIVEN:  I upload a sample xlsx containing a Samples worksheet with a duplicated headers
        WHEN:  we validate the worksheet
        THEN: a ValidationError is thrown

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            ws.append(("Species", db_to_nice_header["external_id"], db_to_nice_header["collection_date"], "", db_to_nice_header["location_id"], db_to_nice_header["note"],
                       db_to_nice_header["location_id"].upper(), " species "))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                spp1_cell = Cell(ws, value="Pf")
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                loc2_cell = Cell(ws, value="Cambodia")
                spp2_cell = Cell(ws, value="Pf")
                row = (spp1_cell, external_id_cell, date_cell,
                       loc_cell, note_cell, loc2_cell, spp2_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="dupHeader_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)

            with pytest.raises(ValidationError) as e:
                samples_iter = validator.validate_manifest_file()
                for sample_dict in samples_iter:
                    if sample_dict and len(sample_dict):
                        pass

            # print ("test_dup_header: " + str(e.exception))
            assert e.value.code == HeaderError.DUP_HEADER
            assert all(c in string.printable for c in str(e.value)), "The Missing Header Validation Error has not been properly interpolated"

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

    def test_case_insensitve_header(self):
        """
        GIVEN:  I upload a valid sample CSV containing mandatory header in different case and trailing/leading spaces than specified in header config ini
        WHEN:  we validate the file
        THEN: the sample validates successfully

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3
            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()

            wb = Workbook()
            ws = wb.create_sheet(title=manifest_config.get_config_sheet_name())

            # Set the header
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       " " + db_to_nice_header["location_id"].upper() + " ", db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                date_cell.number_format = "yyyy-mm-dd"
                loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="caseInsenStrip_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
            samples_iter = validator.validate_manifest_file()

            actual_data = []
            for resp in samples_iter:
                actual_data.append(resp["sample"])
            # NB:  if there is no value for the columns, then it won't be included in the return dict
            expected_data = [{"external_id": "Sample0",
                              "location_id": self.location.location_id,
                              "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample1",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)},
                             {"external_id": "Sample2",
                                 "location_id": self.location.location_id,
                                 "collection_date": datetime.date(year=2000, month=1, day=19)}
                             ]

            assert actual_data == expected_data

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)

# Pending decision on how to implement 150
    def test_formula_or_function(self):
        """
        GIVEN:  I upload a sample xlsx containing a Samples worksheet with a formula or function
        WHEN:  we validate the worksheet
        THEN: we get a ValidationError with an InvalidField code

        """
        workbook_fname = None
        try:

            TOTAL_ROWS = 3

            manifest_config_ini = helpers.DEFAULT_MANIFEST_CONFIG_INI
            manifest_config = BulkUploadExcelConfig(
                header_config_file=manifest_config_ini)
            nice_headers, db_to_nice_header, nice_to_db_header = manifest_config.read_header_config()
            config_sheet_name = manifest_config.get_config_sheet_name()

            # There appears to be a difference in functionality of write only vs
            # read-write workbooks in that read-write workbooks write out the dimensions to the sheet xml.
            wb = Workbook()
            ws = wb.create_sheet(title=config_sheet_name)

            # Set the header
            ws.append((db_to_nice_header["external_id"], db_to_nice_header["collection_date"],
                       db_to_nice_header["location_id"], db_to_nice_header["note"]))

            # Fill in the rows
            for irow in range(TOTAL_ROWS):
                external_id_cell = Cell(ws, value="Sample" + str(irow))
                date_cell = Cell(ws, value=datetime.date(
                    year=2000, month=1, day=19))
                if irow > 0:
                    loc_cell = Cell(ws, value=None)
                    loc_cell.set_explicit_value(value='=C0',
                                                data_type=TYPE_FORMULA)
                    print('Set formula in row '+str(irow))
                else:
                    loc_cell = Cell(ws, value=self.location.curated_name)
                note_cell = Cell(ws, value=None)
                note_cell.set_explicit_value(value='=A0',
                                             data_type=TYPE_FORMULA)
                row = (external_id_cell, date_cell, loc_cell, note_cell)

                ws.append(row)

            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="formula_SampleManifest", delete=False)
            tmp.close()
            workbook_fname = tmp.name
            wb.save(tmp.name)

            total_samples = 0
            validator = BulkAddSamplesValidatorFactory.create_validator(manifest_file=workbook_fname,
                                                                        manifest_file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                                        study=self.study, sample_lookup_func=self.lookup_sample_func)
#            with self.assertRaises(ValidationError) as e:

            loc_id = None
            for sample_dict in validator.validate_manifest_file():
                # Can't actual test this as formula are not resolved on saving with openpyxl
                # https://stackoverflow.com/questions/42376662/openpyxl-2-4-2-cell-value-generated-by-formula-empty-after-saving
                # https://stackoverflow.com/questions/22451973/calculating-excel-sheets-without-opening-them-openpyxl-or-xlwt
                # Could look at https://github.com/anthill/koala
                # if loc_id:
                #     self.assertIn('location_id', sample_dict['sample'],
                #                   'No location_id in validated sample')
                #     self.assertEqual(loc_id, sample_dict['sample']['location_id'], 'Formula not resolved')
                # else:
                #     loc_id = sample_dict['sample']['location_id']
                if sample_dict and sample_dict["valid"]:
                    total_samples += 1

# If data_only=False for load_workbook
#            self.assertEquals(e.exception.code, HeaderError.FORMULA,
#                              "Expected invalid field exception for formula but got " + str(e.exception))
#            self.assertTrue(all(c in string.printable for c in str(e.exception)),
#                            msg="The Invalid Field Validation Error has not been properly interpolated")
            assert total_samples == TOTAL_ROWS, "Expected all samples but got " + str(total_samples)

        finally:
            if workbook_fname and os.path.exists(workbook_fname):
                print("Deleting test file " + str(workbook_fname))
                os.remove(workbook_fname)
